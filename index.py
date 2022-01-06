import logging
import os
import sys

import requests
from openpyxl import load_workbook
from qcloud_cos_v5 import CosConfig, CosServiceError
from qcloud_cos_v5 import CosS3Client

appid = '1307427535'

# 日志配置
logging.basicConfig(level=logging.INFO, stream=sys.stdout)
logger = logging.getLogger()
logger.setLevel(level=logging.INFO)


def main_handler(event, context):
    print(event)

    record = event["Records"][0]
    bucket = record['cos']['cosBucket']['name'] + '-' + str(appid)
    key = record['cos']['cosObject']['key']
    key = key.replace('/' + str(appid) + '/' + record['cos']['cosBucket']['name'] + '/', '', 1)
    logger.info("Get from [%s] to download file [%s]" % (bucket, key))

    secret_id = os.environ.get('TENCENTCLOUD_SECRETID')
    secret_key = os.environ.get('TENCENTCLOUD_SECRETKEY')
    token = os.environ.get('TENCENTCLOUD_SESSIONTOKEN')
    region = os.environ.get('TENCENTCLOUD_REGION')
    config = CosConfig(Secret_id=secret_id, Secret_key=secret_key, Region=region, Token=token)
    client = CosS3Client(config)
    download_path = '/tmp/{}'.format(key)
    old_xlsx = '/tmp/old.xlsx'
    folder = os.path.dirname(download_path)
    if not os.path.exists(folder):
        os.makedirs(folder)

    try:
        response = client.get_object(Bucket=bucket, Key=key, )
        old_response = client.get_object(Bucket=bucket, Key='/scf/old.xlsx', )
        response['Body'].get_stream_to_file(download_path)
        old_response['Body'].get_stream_to_file(old_xlsx)
    except CosServiceError as e:
        print(e.get_error_code())
        print(e.get_error_msg())
        print(e.get_resource_location())
        return "Download Fail"

    # 1.打开 Excel 表格并获取表格名称
    workbook = load_workbook(filename=download_path)
    old_workbook = load_workbook(filename=old_xlsx)
    # 2.通过 sheet 名称获取表格
    sheet = workbook["SCF 整体推进需求"]
    old_sheet = old_workbook["SCF 整体推进需求"]
    contents = []
    template = '''
            {
                "msgtype": "markdown",
                "markdown": {
                    "content": "**SCF需求进度更新**\n需求【%s】的状态变更为：【%s】，产品FO是：%s。"
                }
            }'''
    for row in range(1, 75):
        name = sheet.cell(row=row, column=2).value
        old_name = old_sheet.cell(row=row, column=2).value
        fo = sheet.cell(row=row, column=5).value
        old_fo = old_sheet.cell(row=row, column=5).value
        status = sheet.cell(row=row, column=8).value
        old_status = old_sheet.cell(row=row, column=8).value
        print(str(name) + str(fo) + str(status))
        print(str(old_name) + str(old_fo) + str(old_status))
        if status == '完成' or status == '发布中':
            if status != old_status:
                contents.append(template % (name, status, fo))
    print('变更的状态个数为：' + str(len(contents)))
    for content in contents:
        print(content)
        r = requests.post('https://qyapi.weixin.qq.com/cgi-bin/webhook/send?key=81ef7bfa-0ff4-4850-86df-a8e272c8cafe',
                          data=content.encode('utf-8'))
        print(r.status_code)
        print(r.content)
    return "success"


if __name__ == '__main__':
    event = {
        'Records': [
            {
                'cos': {
                    'cosBucket': {
                        'name': 'woody-chengdu'
                    },
                    'cosObject': {
                        'key': 'scf/new.xlsx'
                    }
                }
            }
        ]
    }
    os.environ.setdefault('TENCENTCLOUD_SECRETID', '')
    os.environ.setdefault('TENCENTCLOUD_SECRETKEY', '')
    os.environ.setdefault('TENCENTCLOUD_REGION', 'ap-chengdu')
    main_handler(event, None)
